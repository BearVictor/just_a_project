import csv
import re
import json
import traceback
import time
import random
import unicodedata
import urllib
import os
import shutil
import difflib
import ast
import threading
import inspect
import hashlib
import os.path
import math
import copy
from lxml import etree, html
from multiprocessing import Pool as ThreadPool
from difflib import SequenceMatcher
from urllib.parse import urlparse
from datetime import datetime, timedelta
from collections import OrderedDict
from inspect import getframeinfo, stack
from shutil import copyfile

try:
    import openpyxl
except:
    pass
try:
    import textract
except:
    pass
try:
    from dictor import dictor
except:
    pass
try:
    import pdf2image
except:
    pass
try:
    import pytesseract
except:
    pass
try:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except:
    pass

global_lock = threading.Lock()


def write_cache(fname, content, *, directory='./cache', ftype='txt', alert=False):
    create_dir(directory)
    fname = str(fname)
    path = '/'.join([directory, hashlib.md5(fname.encode('utf-8')).hexdigest()])
    fpath = '.'.join([path, ftype])
    if ftype == 'txt':
        with open(fpath, 'w', encoding="utf-8") as f:
            f.write(content.replace('&nbsp;', ' '))
    else:
        with open(fpath, 'wb') as f:
            f.write(content)

    if alert:
        print('%s(%s): %s' % (inspect.stack()[0][3], ftype, fname))


def read_cache(fname, directory='./cache', ftype='txt'):
    fname = str(fname)
    path = '/'.join([directory, hashlib.md5(fname.encode('utf-8')).hexdigest()])
    fpath = '.'.join([path, ftype])
    if os.path.exists(fpath):
        if ftype == 'txt':
            with open(fpath, 'r', encoding="utf-8") as f:
                content = f.read()
        else:
            with open(fpath, 'rb') as f:
                content = f.read()

        print('[*] %s: %s' % (inspect.stack()[0][3], fname))
        return content


def delete_cache(fname, directory='./cache', ftype='txt'):
    fname = str(fname)
    path = '/'.join([directory, hashlib.md5(fname.encode('utf-8')).hexdigest()])
    fpath = '.'.join([path, ftype])
    if os.path.exists(fpath):
        os.remove(fpath)
        print('[*] %s(deleted): %s' % (inspect.stack()[0][3], fname))
    else:
        print('[*] %s(notfound): %s ' % (inspect.stack()[0][3], fname))

# --------------------------------------------------
# [XPATH Selenium]
# --------------------------------------------------

def drive_str(driver, XPATH, wait=9):
    ele = WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    return ele.text.strip()

def drive_attribute(driver, XPATH, attribute_name, wait=9):
    ele = WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    return ele.get_attribute(attribute_name)

def drive_find(driver, XPATH, wait=9):
    return WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))

def drive_finds(driver, XPATH, wait=9):
    return WebDriverWait(driver, wait).until(EC.presence_of_all_elements_located((By.XPATH, XPATH)))

def drive_click(driver, XPATH, wait=9, sleep=0.5):
    ele = WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    ele.click()
    time.sleep(sleep)
    return ele

def drive_scroll(driver, XPATH, wait=9, sleep=0.5):
    ele = WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    driver.execute_script("return arguments[0].scrollIntoView(true);", ele)
    time.sleep(sleep)

def drive_innerHTML(driver, XPATH, wait=9):
    ele = WebDriverWait(driver, wait).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    return ele.get_attribute('innerHTML')

def drive_accept_alert(driver, wait=9, sleep=0.5):
    WebDriverWait(driver, wait).until(EC.alert_is_present())
    alert = driver.switch_to_alert()
    alert.accept()
    time.sleep(sleep)

def drive_dismiss_alert(driver, wait=9, sleep=0.5):
    WebDriverWait(driver, wait).until(EC.alert_is_present())
    alert = driver.switch_to_alert()
    alert.dismiss()
    time.sleep(sleep)

def drive_cookies(driver):
    cookies_list = driver.get_cookies()
    cookies_dict = {}
    for cookie in cookies_list:
        cookies_dict[cookie['name']] = cookie['value']
    return cookies_dict
# --------------------------------------------------
# [XPATH request]
# --------------------------------------------------

def xpath_listing(parser, *XPATHs):
    for XPATH in XPATHs:
        arr = parser.xpath(XPATH)
        if arr:
            return arr
    return []

def xpath_string(parser, *XPATHs):
    for XPATH in XPATHs:
        arr = parser.xpath(XPATH)
        if arr:
            string = ' '.join(arr).strip()
            if string:
                return string
    return ''

def xpath_strings(parser, *XPATHs):
    for XPATH in XPATHs:
        arr = parser.xpath(XPATH)
        if arr:
            strings = []
            for string in arr:
                string = string.strip()
                if string:
                    strings.append(string)
            return strings
    return []

def xpath_inner_html(parser, *XPATHs):
    for XPATH in XPATHs:
        eles = parser.xpath(XPATH)
        if eles:
            ele_str = etree.tostring(eles[0])
            return ele_str.decode("utf-8")
    return ''

def inner_html_ele(ele_parser):
    ele_str = etree.tostring(ele_parser)
    return ele_str.decode("utf-8")

def get_number_fstring(str):
    number = re.search(r'\d[.,\d]*\d+|\d+', str)
    if number.group():
        clear_number = number.group().replace(',','')
        return clear_number

def get_numbers_fstring(str):
    numbers = re.findall(r'\d[.,\d]*\d+|\d+', str)
    if numbers:
        clear_numbers = [number.replace(',','') for number in numbers]
        return clear_numbers

def get_money_fstring(str):
    return re.findall(r'(\d[\d.,]*)\b', str)[0] if str else ''

def get_emails_fstring(str):
    return re.findall(r'[\w\.-]+@[\w\.-]+\.\w+', str)

def get_urls_fstring(text):
    return re.findall("(?:(?:https?|ftp):\/\/)?[\w/\-?=%.]+\.[\w/\-?=%]+", text)

def encode_url(url):
    return urllib.parse.quote(url, safe='')

def pool(function_name, parameter_list, processes=3):
    pool = ThreadPool(processes)
    pool.map(function_name, parameter_list)
    pool.terminate()
    pool.join()

def write_dicts_csv(data, fname):
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)
    print('[*] %s(len): %s' % (inspect.stack()[0][3], len(data)))

    fname = re.sub(r'\.\w{2,}$', '', fname) + '.csv'
    with open(fname, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        for i in data:
            writer.writerow(i)

    print(' [+] %s(done): %s' % (inspect.stack()[0][3], fname))

def write_dicts_xlsx(dict_list, fname):
    print(f'{inspect.stack()[0][3]} (len):', len(dict_list))
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)

    path = re.sub(r'\.\w{2,}$', '', fname) + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active

    if dict_list:
        field_names = list(dict_list[0].keys())
        ws.append(field_names)

        for dict in dict_list:
            try:
                values = (dict[k] for k in field_names)
                ws.append(values)
            except:
                print(dict)
                raise
    wb.save(path)
    print("Writing %s Finished!" % fname)

def write_dicts_xlsx_gen(dicts_gen, fname):
    print(f'{inspect.stack()[0][3]}():', 'is writting ...')
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)

    path = re.sub(r'\.\w{2,}$', '', fname) + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active

    rows = 0
    if dicts_gen:
        field_names = []
        for index, dict in enumerate(dicts_gen):
            if index == 0:
                field_names = list(dict.keys())
                ws.append(field_names)

                values = [dict[k] for k in field_names]
                ws.append(values)
            else:
                try:
                    values = [dict[k] for k in field_names]
                    ws.append(values)
                except:
                    print(dict)
                    raise
            rows += 1
    wb.save(path)
    print("Writing %s (%s rows) Finished!" % (fname, rows))

def write_json(data, fname):
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.json'
    with open('%s.json' % fname, 'w') as jsonfile:
        json.dump(data, jsonfile, indent=4)
        print("Writing %s.json Finished!" % fname)

def write_txt(text='', fname ='raw', alert=True):
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with open(fname, 'w', encoding="utf-8") as f:
        f.write(text)
        if alert:
            caller = getframeinfo(stack()[1][0])
            caller_fname = re.sub(r'^.*/', '', caller.filename)
            caller_line = caller.lineno
            print("[*] %s(%s l%s): (%s) %s..."
                  % (inspect.stack()[0][3], caller_fname, caller_line, fname, text.encode()[0:135]))

def write_html(text='', fname ='raw', alert=True):
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.html'
    with open(fname, 'w') as f:
        f.write(text)
        if alert:
            caller = getframeinfo(stack()[1][0])
            caller_fname = re.sub(r'^.*/', '', caller.filename)
            caller_line = caller.lineno
            print("[*] %s(%s l%s): (%s) %s..."
                  % (inspect.stack()[0][3], caller_fname, caller_line, fname, text.encode()[0:135]))

def append_txt(text='', fname ='raw', alert=True):
    dir_path = os.path.dirname(fname)
    if dir_path:
        create_dir(dir_path)
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with global_lock:
        with open(fname, 'a', encoding="utf-8") as f:
            f.write(text + '\n')
            if alert:
                message('%s(%s)' % (inspect.stack()[0][3], fname), text)

def list_txt(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    lines = read_txt_lines(fname)
    first_data = json.loads(lines[0])
    if isinstance(first_data, dict):
        return [json.loads(i) for i in lines]
    else:
        pdatas = []
        for i in lines:
            pdatas.extend(json.loads(i))
        return pdatas

def list_txt_gen(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with open(fname, 'r', encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines:
            line_data = json.loads(line)
            if isinstance(line_data, dict):
                yield line_data
            else:
                for item in line_data:
                    yield item

def read_json(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.json'
    with open(fname) as f:
        return json.load(f)

def read_dicts_cvs(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.csv'
    results = []
    with open(fname, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        headers = []
        for row_index, row in enumerate(reader):
            if row_index == 0:
                headers = row
            else:
                adding = {}
                for header_index, header in enumerate(headers):
                    adding[header] = row[header_index]
                results.append(adding)
        return results

def read_dicts_cvs_latin(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.csv'
    results = []
    with open(fname, 'r', encoding='ISO-8859-1', newline='') as f:
        reader = csv.reader(f)
        headers = []
        for row_index, row in enumerate(reader):
            if row_index == 0:
                headers = row
            else:
                adding = {}
                for header_index, header in enumerate(headers):
                    adding[header] = row[header_index]
                results.append(adding)
        return results

def read_dicts_xlsx(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.xlsx'
    wb = openpyxl.load_workbook(fname)
    ws = wb.worksheets[0]

    header = [cell.value.strip() for cell in ws[1]]

    dicts = []
    for row_index, row in enumerate(ws.iter_rows()):
        if row_index == 0:
            continue
        values = {}
        for key, cell in zip(header, row):
            cell_val = '' if cell.value is None else cell.value
            values[key] = cell_val
        dicts.append(values)
    return dicts

def read_txt(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with open(fname, 'r', encoding="utf-8") as f:
        return f.read()

def read_txt_lines(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with open(fname, 'r', encoding="utf-8") as f:
        lines = f.readlines()
        new_lines = [line.strip() for line in lines if line.strip()]
        return new_lines

def read_txt_lines_gen(fname):
    fname = re.sub(r'\.\w{2,}$', '', fname) + '.txt'
    with open(fname, 'r', encoding="utf-8") as f:
        lines = f.readlines()
        new_lines = (line.strip() for line in lines)
        return new_lines

def clear_string(input_string):
    clear_string = input_string.strip()
    clear_string = re.sub('[^a-zA-Z0-9 \n\.\-@]', '', clear_string)
    return clear_string

def remove_words(input_str, removing_words=[]):
    for x in removing_words:
        input_str = input_str.replace(x, '').strip()
    return input_str

def clear_name_file(name):
    name = unicodedata.normalize('NFD', name).encode('ascii', 'ignore')
    name = name.decode("utf-8")
    name = clear_string(name)
    name = ' '.join(name.split())
    return name

def similarity(input, match):
    lower_input = input.lower()
    lower_match = match.lower()
    return round(SequenceMatcher(None, lower_input, lower_match).ratio(), 3)

def get_closests(input, matches):
    unique_matches = list(OrderedDict.fromkeys(matches))
    lower_matches = [x.lower() for x in unique_matches]
    lower_input = input.lower()

    close_matches = difflib.get_close_matches(lower_input, lower_matches)

    closests = []
    for x in close_matches:
        index = lower_matches.index(x)
        closests.append(unique_matches[index])
    return closests

def waitToTomorrow():
    """Wait to tommorow 00:00 am"""
    tomorrow = datetime.replace(datetime.now() + timedelta(days=1), hour=0, minute=0, second=0)
    delta = tomorrow - datetime.now()
    time.sleep(delta.seconds)


def is_weekend(date_input=datetime.today()):
    """ is weekend
    The day of the week as an integer, where Monday is 0 and Sunday is 6.
    :return: True Or False
    """
    weekno = date_input.weekday()
    if weekno >= 5:
        return True
    return False


def sleep(sl_time=45, message=''):
    caller = getframeinfo(stack()[1][0])
    funcname = re.sub(r'^.*/', '', caller.filename)
    print("[*] sleep(%s l%d): (%ss) %s" % (funcname, caller.lineno, sl_time, message))
    time.sleep(sl_time)

def random_chome_useragent():
    useragent_lines = """
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
        Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36
        Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36
        Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
        Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36
        Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36
        Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36
        Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.143 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36
        Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36
    """
    useragents = useragent_lines.split('\n')
    useragents = [i.strip() for i in useragents if i.strip()]
    return random.choice(useragents)

def random_firefox_useragent():
    useragent_lines = """
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:18.0) Gecko/20100101 Firefox/18.0
        Mozilla/5.0 (Windows NT 5.1; rv:36.0) Gecko/20100101 Firefox/36.0
        Mozilla/5.0 (Windows NT 5.1; rv:33.0) Gecko/20100101 Firefox/33.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0
        Mozilla/5.0 (Windows NT 10.0; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0
        Mozilla/5.0 (Windows NT 10.0; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:67.0) Gecko/20100101 Firefox/67.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:17.0) Gecko/20100101 Firefox/17.0
        Mozilla/5.0 (Windows NT 6.0; rv:34.0) Gecko/20100101 Firefox/34.0
        Mozilla/5.0 (Windows NT 10.0; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0
        Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0
        Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:24.0) Gecko/20100101 Firefox/24.0
        Mozilla/5.0 (Windows NT 5.1; rv:40.0) Gecko/20100101 Firefox/40.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0
        Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:73.0) Gecko/20100101 Firefox/73.0
        Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0
        Mozilla/5.0 (Windows NT 6.1; rv:17.0) Gecko/20100101 Firefox/20.6.14
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0
        Mozilla/5.0 (Windows NT 5.1; rv:30.0) Gecko/20100101 Firefox/30.0
        Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10; rv:33.0) Gecko/20100101 Firefox/33.0
        Mozilla/5.0 (Windows NT 6.1; rv:52.0) Gecko/20100101 Firefox/52.0
        Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:25.0) Gecko/20100101 Firefox/29.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:47.0) Gecko/20100101 Firefox/47.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:62.0) Gecko/20100101 Firefox/62.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
        Mozilla/5.0 (Windows NT 6.1; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0
        Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
        Mozilla/5.0 (Windows NT 6.3; WOW64; rv:57.0) Gecko/20100101 Firefox/57.0
        Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0
        Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:73.0) Gecko/20100101 Firefox/73.0
        Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
        Mozilla/5.0 (Windows NT 5.1; rv:52.0) Gecko/20100101 Firefox/52.0
    """
    useragents = useragent_lines.split('\n')
    useragents = [i.strip() for i in useragents if i.strip()]
    return random.choice(useragents)

def random_useragent():
    useragent_lines = """
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
            Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36
            Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36
            Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
            Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36
            Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36
            Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.143 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36
            Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:18.0) Gecko/20100101 Firefox/18.0
            Mozilla/5.0 (Windows NT 5.1; rv:36.0) Gecko/20100101 Firefox/36.0
            Mozilla/5.0 (Windows NT 5.1; rv:33.0) Gecko/20100101 Firefox/33.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0
            Mozilla/5.0 (Windows NT 10.0; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0
            Mozilla/5.0 (Windows NT 10.0; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:67.0) Gecko/20100101 Firefox/67.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:17.0) Gecko/20100101 Firefox/17.0
            Mozilla/5.0 (Windows NT 6.0; rv:34.0) Gecko/20100101 Firefox/34.0
            Mozilla/5.0 (Windows NT 10.0; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0
            Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0
            Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:24.0) Gecko/20100101 Firefox/24.0
            Mozilla/5.0 (Windows NT 5.1; rv:40.0) Gecko/20100101 Firefox/40.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0
            Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:73.0) Gecko/20100101 Firefox/73.0
            Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0
            Mozilla/5.0 (Windows NT 6.1; rv:17.0) Gecko/20100101 Firefox/20.6.14
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0
            Mozilla/5.0 (Windows NT 5.1; rv:30.0) Gecko/20100101 Firefox/30.0
            Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10; rv:33.0) Gecko/20100101 Firefox/33.0
            Mozilla/5.0 (Windows NT 6.1; rv:52.0) Gecko/20100101 Firefox/52.0
            Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:25.0) Gecko/20100101 Firefox/29.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:47.0) Gecko/20100101 Firefox/47.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:62.0) Gecko/20100101 Firefox/62.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:45.0) Gecko/20100101 Firefox/45.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
            Mozilla/5.0 (Windows NT 6.1; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0
            Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
            Mozilla/5.0 (Windows NT 6.3; WOW64; rv:57.0) Gecko/20100101 Firefox/57.0
            Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0
            Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:73.0) Gecko/20100101 Firefox/73.0
            Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:58.0) Gecko/20100101 Firefox/58.0
            Mozilla/5.0 (Windows NT 5.1; rv:52.0) Gecko/20100101 Firefox/52.0
        """
    useragents = useragent_lines.split('\n')
    useragents = [i.strip() for i in useragents if i.strip()]
    return random.choice(useragents)

def get_host_name(url):
    parsed_uri = urlparse(url)
    return '{uri.scheme}://{uri.netloc}'.format(uri=parsed_uri)

# -------------------------------------------------------------
# [get_host_post_user_pass, get_random_proxy_cookies, get_proxies]
# -------------------------------------------------------------


def extract_proxies(fpath):
    proxies = [line.strip() for line in read_txt_lines(fpath) if '#' not in line]
    return proxies

def extract_proxies_au(fpath, regex_compile=re.compile(r'(?P<host>.+):(?P<port>.+):(?P<user>.+):(?P<pass>.+)')):
    proxies = [regex_compile.sub(r'\g<user>:\g<pass>@\g<host>:\g<port>', line).strip() for line in read_txt_lines(fpath) if '#' not in line]
    return proxies

def extract_host_port_user_pass(proxy, regex_compile=re.compile(r'(?P<user>.+):(?P<pass>.+)@(?P<host>.+):(?P<port>.+)')):
    proxy = regex_compile.search(proxy)
    return proxy.group('host'), proxy.group('port'), proxy.group('user'), proxy.group('pass')


def unique_list(l):
    return list(OrderedDict.fromkeys(l))

def unique_dicts(dicts):
    seen = set()
    unique_dicts = []
    for d in dicts:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            unique_dicts.append(d)
    return unique_dicts

def unique_dicts_by_key(dicts, key):
    unique_keys = []
    unique_dicts = []
    for i in dicts:
        if i[key] not in unique_keys:
            unique_keys.append(i[key])
            unique_dicts.append(i)
    return unique_dicts


def to_str(bytes_or_str):
    return bytes_or_str.decode('utf-8') if isinstance(bytes_or_str, bytes) else bytes_or_str


def to_bytes(bytes_or_str):
    return bytes_or_str.encode('utf-8') if isinstance(bytes_or_str, str) else bytes_or_str


def sort_dicts(input_list, key, reverse=False):
    """Sort dicts by key"""
    input_list.sort(key=lambda item: item[key], reverse=reverse)
    return input_list

# -------------------------------------------------------------
# [delete_file, create_directory, delete_directory]
# -------------------------------------------------------------
def delete_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print("delete_file %s!" % file_path)
    return file_path

def create_dir(directory, *, alert=True):
    if not os.path.exists(directory):
        os.makedirs(directory)
        if alert:
            print('[*] %s: %s' % (inspect.stack()[0][3], directory))
    return directory

def delete_dir(directory):
    if os.path.exists(directory):
        shutil.rmtree(directory)
        print('delete_directory: %s!' % directory)
    return directory


# -------------------------------------------------------------
# [return_literal]
# -------------------------------------------------------------

def literal(input):
    try:
        return ast.literal_eval(input)
    except:
        return input

# -------------------------------------------------------------
# [alert]
# -------------------------------------------------------------
def message(title, content='', *, maxlen=135):
    """ Print the messages
    :param title: title
    :param content: content
    :return:
    """
    if content:
        content = json.dumps(content)
        content = content[:maxlen] + '...' if len(content) > maxlen else content
        print("%s: %s" % (title, content))
    else:
        print("%s" % title)

def lister(l, index, exception=''):
    try:
        return l[index]
    except:
        return exception

def dictor(d, key, exception=''):
    try:
        return d[key]
    except:
        return exception

def debug_info(message='debug_info'):
    caller = getframeinfo(stack()[1][0])
    print("%s:%d - %s" % (caller.filename, caller.lineno, message))

def move_src_to_dst(src, dst):
    """ Move files of src to dest """
    src_files = os.listdir(src)
    create_dir(dst)
    for src_file in src_files:
        src_path = src + '/' + src_file
        dst_path = dst + '/' + src_file
        shutil.move(src_path, dst_path)
        print('%s -> %s' % (src_path, dst_path))
    delete_dir(src)


def encode_ascii(string):
    """ ex: © -> &#169; """
    return to_str(string.encode(encoding="ascii", errors="xmlcharrefreplace"))


def add_host(host_url, url):
    host_name = get_host_name(host_url)
    if not url:
        return url
    if re.search(r'(https?|:|//|www)', url):
        return url
    if host_name not in url:
        return host_name + '/' + re.sub(r'^/', '', url)


def timing(f):
    def wrap(*args):
        start = datetime.now()
        ret = f(*args)
        print(f'{f.__name__}() took {datetime.now() - start}')
        return ret
    return wrap

def re_emails(text, emailRegex=None):
    if emailRegex is None:
        emailRegex = re.compile(r'''(
            [a-zA-Z0-9._%+-]+       # username
            @                       # @ symbol
            [a-zA-Z0-9.-]+          # domain name
            (\.[a-zA-Z]{2,4})       # dot-something
            )''', re.VERBOSE)

    matches = []
    for groups in emailRegex.findall(text):
        matches.append(groups[0])
    return matches

def re_phones(text, phoneRegex=None):
    if phoneRegex is None:
        phoneRegex = re.compile(r'''(
            (\d{3}|\(\d{3}\))       # area code
            (\s|-|\.)?              # separator
            (\d{3})                 # first 3 digits
            (\s|-|\.)?              # separator
            (\d{4})                 # last 4 digits
            (\s*(ext|x|ext.)\s*(\d{2,5}))?  # extension
            )''', re.VERBOSE)

    matches = []
    for groups in phoneRegex.findall(text):
        matches.append(groups[0])
    return matches

def htmltotext(html_text):
    myDict = {
        re.compile(r'>\s+<'): '><',  # remove space between tags
        re.compile(r'<!--.*?(-->)', re.I | re.DOTALL): '\n',  # remove comment
        re.compile(r'<(/p|/h\d|/li|br/?)>', re.I): '\n',  # replace new lines
        re.compile(r'<[^>]*>'): '',
        re.compile(r' +\.'): '.',
        re.compile(r' +,'): ',',
        re.compile(r'  +'): ' ',
    }
    for rx, repl in myDict.items():
        html_text = rx.sub(repl, html_text)

    parser = html.fromstring('<p>%s</p>' % html_text)
    text = xpath_string(parser, './text()')
    return text.strip()

def multiple_resub(myDict, text):
    """
    myDict = {
        re.compile(r' \(\/ws.+\)'): '',
        re.compile(r'Compiled on.{36}'): '',
        re.compile(re.escape(r' IronWare')): 'IronWare'
    }
    """
    for rx,repl in myDict.items():
        text = rx.sub(repl, text)
    return text


def extract_pdf_image_text(fpath):
    """
    pip install textract
    """
    return textract.process(fpath, method='tesseract', language='eng', encoding='utf-8').decode('utf-8')

def extract_scanded_pdf(pdf_path, value_regexs):
    images = pdf2image.convert_from_path(pdf_path)
    pdf_text = ''
    for image in images:
        pil_im = image
        ocr_dict = pytesseract.image_to_data(pil_im, lang='eng', output_type=pytesseract.Output.DICT)

        txt = " ".join(ocr_dict['text'])
        for value_re in value_regexs:
            value_found = value_re.search(txt)
            if value_found:
                pdf_text += '\n%s' % txt
                print('[*] %s(value_found): %s' % (inspect.stack()[0][3], value_found.group()))
                return pdf_text

        pdf_text += '\n%s' % txt
    return pdf_text

def hash_name(name):
    return hashlib.md5(name.encode('utf-8')).hexdigest()

def lines(doc):
     return [i.strip() for i in doc.split('\n') if i.strip()]

def zipdir(zip_path, src_dir):
    shutil.make_archive(zip_path, 'zip', src_dir)
    print('[*] %s(done): %s' % (inspect.stack()[0][3], src_dir))

def log_erros(func):
    log_fname = inspect.stack()[0][3]
    def wrap(*args, **kwargs):
        try:
            func(*args, **kwargs)
        except:
            log_dir = create_dir('./log')
            e_message = traceback.format_exc()
            exception_fname = '%s/%s_%s' % (log_dir, func.__name__, datetime.now().strftime('%y_%m_%d'))

            write_txt(e_message, exception_fname, alert=False)
            print('[*] %s(done): %s' % (log_fname, exception_fname))
    return wrap

def log_erros_with_del_cache(func):
    log_fname = inspect.stack()[0][3]
    def wrap(*args, **kwargs):
        try:
            func(*args, **kwargs)
        except:
            log_dir = create_dir('./log')
            e_message = traceback.format_exc()
            exception_fname = '%s/%s_%s' % (log_dir, func.__name__, datetime.now().strftime('%y_%m_%d'))

            # todo: delete cache dir
            cache_dir = kwargs.get('cache_dir', '')
            if cache_dir:
                delete_dir(cache_dir)

            # todo: write erros
            write_txt(e_message, exception_fname, alert=False)
            print('[*] %s(done): %s' % (log_fname, exception_fname))
    return wrap

def copy_dict(data, exclusions=['url', 'parent_url']):
    new_dict = copy.deepcopy(data)
    for key in exclusions:
        new_dict.pop(key, '')
    return new_dict

def rename_dir(path, new_path):
    if os.path.exists(path):
        delete_dir(new_path)
        os.rename(path, new_path)
        print('rename_dir(done): %s -> %s' % (path, new_path))

def try_excetp_re(re_found, group_name):
    try:
        group_value = re_found.group(group_name).strip()
    except:
        group_value = ''
    return group_value

def deleteFiles(dir_name, extension):
    for item in os.listdir(dir_name):
        if item.endswith(extension):
            print(os.path.join(dir_name, item))
            os.remove(os.path.join(dir_name, item))

def chunks(lst, n):
    return [lst[i:i + n] for i in range(0, len(lst), n)]

def chunks_gen(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def split_list(alist, wanted_parts=1):
    length = len(alist)
    return [alist[i * length // wanted_parts: (i + 1) * length // wanted_parts]
            for i in range(wanted_parts)]

def separate_xlsx(fname, parts_number):
    fname = re.sub(r'\.\w{2,}$', '', fname)
    alist = read_dicts_xlsx(fname)
    parts = split_list(alist, parts_number)
    for index, l in enumerate(parts):
        part_len = len(l)
        part_fname = f'{fname}-part{index + 1}-{part_len}records'
        write_dicts_xlsx(l, part_fname)

def chunk_xlsx(fname, step):
    fname = re.sub(r'\.\w{2,}$', '', fname)
    alist = read_dicts_xlsx(fname)
    parts = chunks(alist, step)
    for index, l in enumerate(parts):
        part_len = len(l)
        part_fname = f'{fname}-part{index + 1}-{part_len}records'
        write_dicts_xlsx(l, part_fname)

def deleteFiles(dir_name, extension):
    for item in os.listdir(dir_name):
        if item.endswith(extension):
            print(os.path.join(dir_name, item))
            os.remove(os.path.join(dir_name, item))

if __name__=="__main__":
    print('Testing....')



